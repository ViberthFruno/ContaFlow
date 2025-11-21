# excel_processor.py
"""
Procesador de Excel y XML para ContaFlow con soporte empresarial, rutas dinámicas automáticas,
filtrado por fecha tanto en XMLs como en filas de Excel, procesamiento especial de PDFs de Correos,
y nueva funcionalidad de extracción de códigos de placas desde el campo <OtroTexto>.
Maneja búsqueda XML con rutas dinámicas basadas en año/mes actual, matching con Excel después de
filtrado por fecha, generación de archivos por empresa incluyendo actividades comerciales
configuradas, campo Paquete y procesamiento especial para facturas de Correos de Costa Rica SA.
"""

import os
import glob
import xml.etree.ElementTree as ET
import threading
import time
import unicodedata
from datetime import datetime
from typing import Dict, List, Tuple, Optional

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Importar el procesador de PDFs de Correos
try:
    from contaflow.processors.pdf_processor import CorreosPDFProcessor, create_correos_pdf_processor

    PDF_PROCESSOR_AVAILABLE = True
except ImportError:
    PDF_PROCESSOR_AVAILABLE = False
    print("Advertencia: pdf_processor no está disponible. Funcionalidad de Correos limitada.")

# Importar el nuevo procesador de OtroTexto para extracción de placas
try:
    from contaflow.processors.otro_texto_processor import create_otro_texto_processor

    OTRO_TEXTO_PROCESSOR_AVAILABLE = True
except ImportError:
    OTRO_TEXTO_PROCESSOR_AVAILABLE = False
    print("Advertencia: otro_texto_processor no está disponible. Funcionalidad de extracción de placas limitada.")


class ExcelProcessor:
    """Clase principal para el procesamiento de archivos Excel con datos XML por empresa, rutas dinámicas, filtrado por fecha en XMLs y Excel, soporte para PDFs de Correos, y extracción de placas desde OtroTexto."""

    def __init__(self, automation_tab=None):
        self.automation_tab = automation_tab

        if not OPENPYXL_AVAILABLE:
            self.log_message("❌ Error: openpyxl no está disponible. Instale con: pip install openpyxl", "error")
            raise ImportError("openpyxl requerido para procesamiento Excel")

        # Inicializar procesador de PDFs de Correos
        self.correos_pdf_processor = None
        if PDF_PROCESSOR_AVAILABLE:
            try:
                self.correos_pdf_processor = create_correos_pdf_processor()
                if self.correos_pdf_processor:
                    self.log_message("✅ Procesador de PDFs de Correos inicializado", "info")
                else:
                    self.log_message("⚠️ No se pudo inicializar procesador de PDFs (pdfplumber faltante)", "warning")
            except Exception as e:
                self.log_message(f"⚠️ Error inicializando procesador de PDFs: {e}", "warning")

        # Inicializar procesador de OtroTexto para extracción de placas
        self.otro_texto_processor = None
        if OTRO_TEXTO_PROCESSOR_AVAILABLE:
            try:
                self.otro_texto_processor = create_otro_texto_processor()
                self.log_message("✅ Procesador de OtroTexto para extracción de placas inicializado", "info")
            except Exception as e:
                self.log_message(f"⚠️ Error inicializando procesador de OtroTexto: {e}", "warning")

        self.company_file_names = {
            'nargallo': 'NargalloDelEste',
            'ventas_fruno': 'VentasFruno',
            'creme_caramel': 'CremeCaramel',
            'su_laka': 'SuLaka'
        }

        self.combustible_exclusion_emitters = set()

        # Columnas de Excel - Incluye "Paquete" antes de "Actividad Comercial"
        self.excel_columns = [
            'Proveedor', 'Numero', 'Tipo de Documento', 'Fecha Documento',
            'Fecha Rige', 'Aplicación', 'Monto', 'Subtotal', 'Impuesto1',
            'Tipo de Cambio', 'Notas', 'Condicion de Pago', 'Moneda',
            'Subtipo Documento', 'Fecha Vence', 'Tipo Asiento', 'Paquete', 'Actividad Comercial'
        ]

        # Índice de la columna "Fecha Documento" (empezando desde 0)
        self.fecha_documento_column_index = 3

        self.red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        self.red_font = Font(color="FFFFFF", bold=True)  # Texto blanco para mejor legibilidad
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        self.stats = {
            'companies_processed': 0, 'companies_with_matches': 0, 'companies_without_matches': 0,
            'total_xml_count': 0, 'total_xml_valid': 0, 'total_xml_current_month': 0,
            'total_xml_excluded_by_date': 0, 'total_matches': 0,
            'total_manual_reviews': 0, 'excel_processed': 0,
            'files_created': 0, 'processing_time': 0, 'company_details': {},
            'companies_no_matches': [], 'excluded_by_date_details': [],
            # Estadísticas para Correos
            'correos_pdfs_processed': 0, 'correos_pdfs_failed': 0, 'correos_matches': 0,
            # Estadísticas para rutas dinámicas
            'companies_folders_found': 0, 'companies_folders_missing': 0, 'companies_folders_skipped': [],
            # Nuevas estadísticas para filtrado de Excel
            'excel_rows_total': 0, 'excel_rows_current_month': 0, 'excel_rows_excluded_by_date': 0,
            'excel_excluded_by_date_details': [],
            # Nuevas estadísticas para extracción de placas desde OtroTexto
            'otro_texto_processed': 0, 'placas_extracted': 0, 'placas_failed': 0,
            'fallback_to_detalle': 0, 'placa_extraction_rate': 0.0,
            # Estadísticas para exclusiones de combustible
            'combustible_exclusions_applied': 0
        }

        self.stop_event = threading.Event()
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year

    def _parse_excel_date(self, fecha_documento_text: str) -> Optional[datetime]:
        """
        Parsea la fecha del documento Excel desde formato dd-mm-yyyy.

        Args:
            fecha_documento_text (str): Fecha en formato '27-06-2024'

        Returns:
            datetime: Objeto datetime o None si no se puede parsear
        """
        try:
            if not fecha_documento_text:
                return None

            # Convertir a string si es necesario
            fecha_str = str(fecha_documento_text).strip()

            # Formato esperado: dd-mm-yyyy (27-06-2024)
            if len(fecha_str) == 10 and fecha_str.count('-') == 2:
                return datetime.strptime(fecha_str, '%d-%m-%Y')
            # Formato alternativo: dd/mm/yyyy
            elif len(fecha_str) == 10 and fecha_str.count('/') == 2:
                return datetime.strptime(fecha_str, '%d/%m/%Y')
            else:
                return None

        except (ValueError, TypeError) as e:
            return None

    def _is_excel_row_current_month(self, fecha_documento: datetime) -> bool:
        """
        Verifica si una fecha del Excel pertenece al mes y año actual.

        Args:
            fecha_documento (datetime): Fecha a verificar

        Returns:
            bool: True si es del mes actual
        """
        return (fecha_documento.month == self.current_month and
                fecha_documento.year == self.current_year)

    def _is_correos_xml(self, xml_file_path: str) -> bool:
        """
        Verifica si un archivo XML corresponde a una factura de Correos de Costa Rica SA.

        Args:
            xml_file_path (str): Ruta del archivo XML

        Returns:
            bool: True si es una factura de Correos
        """
        try:
            tree = ET.parse(xml_file_path)
            root = tree.getroot()

            # Buscar elemento Nombre que contenga exactamente "Correos de Costa Rica SA"
            for elem in root.iter():
                if 'Nombre' in elem.tag and elem.text:
                    nombre_text = elem.text.strip()
                    if nombre_text == "Correos de Costa Rica SA":
                        self.log_message(f"📮 Detectada factura de Correos: {os.path.basename(xml_file_path)}", "info")
                        return True

            return False

        except Exception as e:
            self.log_message(f"⚠️ Error verificando XML de Correos {xml_file_path}: {e}", "warning")
            return False

    def _process_correos_xml(self, xml_file_path: str, numero_consecutivo: str) -> Tuple[bool, Optional[str]]:
        """
        Procesa un XML de Correos usando el PDF asociado.

        Args:
            xml_file_path (str): Ruta del archivo XML
            numero_consecutivo (str): Número consecutivo del XML

        Returns:
            Tuple[bool, str]: (éxito, dato_procesado)
        """
        try:
            if not self.correos_pdf_processor:
                self.log_message("❌ Procesador de PDFs no disponible para factura de Correos", "error")
                return False, None

            self.log_message(f"📮 Procesando factura de Correos: {numero_consecutivo}", "info")

            # Procesar PDF usando el procesador especializado
            success, formatted_data = self.correos_pdf_processor.process_correos_pdf(xml_file_path)

            if success and formatted_data:
                self.stats['correos_pdfs_processed'] += 1
                self.stats['correos_matches'] += 1
                self.log_message(f"✅ PDF de Correos procesado: {formatted_data[:80]}...", "success")
                return True, formatted_data
            else:
                self.stats['correos_pdfs_failed'] += 1
                self.log_message(f"❌ No se pudo procesar PDF de Correos para {numero_consecutivo}", "error")
                return False, None

        except Exception as e:
            self.stats['correos_pdfs_failed'] += 1
            self.log_message(f"❌ Error procesando XML de Correos {numero_consecutivo}: {e}", "error")
            return False, None

    def _apply_combustible_exclusion_if_needed(self, root, numero: str, xml_data: Dict[str, List[str]],
                                               company_stats: Dict) -> bool:
        """Aplica la exclusión de combustible si el emisor está configurado."""
        should_skip, emisor_name = self._should_skip_combustible_extraction(root)

        if not should_skip:
            return False

        detalles = self._extract_detalle_list(root)
        xml_data[numero] = detalles
        company_stats['combustible_exclusions'] += 1
        self.stats['combustible_exclusions_applied'] += 1

        if emisor_name:
            self.log_message(f"⛽️ Emisor excluido de placas: {emisor_name}", "info")
        else:
            self.log_message("⛽️ Se aplicó exclusión de combustible (NombreEmisor no disponible)", "info")

        return True

    def _extract_detalle_list(self, root) -> List[str]:
        """Obtiene la lista de detalles desde el XML."""
        try:
            return [
                elem.text.strip()
                for elem in root.iter()
                if 'Detalle' in elem.tag and elem.text and elem.text.strip()
            ]
        except Exception:
            return []

    def _should_skip_combustible_extraction(self, root) -> Tuple[bool, Optional[str]]:
        """Determina si se debe omitir la extracción de placa por exclusión configurada."""
        if not self.combustible_exclusion_emitters:
            return False, None

        emisor_name = self._extract_emisor_name(root)
        if not emisor_name:
            return False, None

        normalized = self._normalize_emisor_name(emisor_name)
        if normalized in self.combustible_exclusion_emitters:
            return True, emisor_name

        return False, emisor_name

    def _extract_emisor_name(self, root) -> Optional[str]:
        """Extrae el NombreEmisor desde el XML."""
        try:
            for elem in root.iter():
                tag = self._strip_namespace(elem.tag)
                if tag == 'NombreEmisor' and elem.text and elem.text.strip():
                    return elem.text.strip()

            for elem in root.iter():
                tag = self._strip_namespace(elem.tag)
                if tag == 'Emisor':
                    for child in elem.iter():
                        child_tag = self._strip_namespace(child.tag)
                        if child_tag in ('NombreEmisor', 'Nombre') and child.text and child.text.strip():
                            return child.text.strip()
                    break
        except Exception:
            return None

        return None

    @staticmethod
    def _strip_namespace(tag: str) -> str:
        if not tag:
            return ''
        if '}' in tag:
            return tag.split('}', 1)[1]
        return tag

    @staticmethod
    def _normalize_emisor_name(name: str) -> str:
        if not name:
            return ''
        normalized = unicodedata.normalize('NFKD', name)
        normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
        return normalized.strip().lower()

    def _extract_otro_texto_info(self, root) -> Optional[str]:
        """
        Extrae información del campo <OtroTexto> y procesa códigos de placas.

        Args:
            root: Elemento raíz del XML

        Returns:
            str: Información procesada del OtroTexto o None si no se encuentra
        """
        try:
            if not self.otro_texto_processor:
                return None

            # Buscar el campo OtroTexto
            otro_texto_content = None
            for elem in root.iter():
                if 'OtroTexto' in elem.tag and elem.text:
                    otro_texto_content = elem.text.strip()
                    break

            if not otro_texto_content:
                return None

            # Procesar el contenido de OtroTexto para extraer placas
            self.stats['otro_texto_processed'] += 1

            formatted_placa = self.otro_texto_processor.process_otro_texto(otro_texto_content)

            if formatted_placa:
                self.stats['placas_extracted'] += 1
                self.log_message(f"🚗 Placa extraída: {formatted_placa}", "info")
                return formatted_placa
            else:
                self.stats['placas_failed'] += 1
                return None

        except Exception as e:
            self.stats['placas_failed'] += 1
            self.log_message(f"⚠️ Error procesando OtroTexto: {e}", "warning")
            return None

    def process_excel_files(self, input_folder: str, xml_folder: str, output_folder: str,
                            config: Optional[Dict] = None) -> Dict:
        """Proceso principal de archivos Excel con XML empresarial, rutas dinámicas, filtrado por fecha en XMLs y Excel, soporte para Correos, y extracción de placas desde OtroTexto."""
        start_time = time.time()
        self.log_message(
            "🚀 Iniciando procesamiento Excel/XML empresarial con rutas dinámicas, filtrado por fecha en XMLs y Excel, soporte para Correos, y extracción de placas desde OtroTexto",
            "info")

        try:
            self._reset_stats()
            self.combustible_exclusion_emitters = set()

            if not config or 'company_folders' not in config:
                return {'success': False, 'error': 'Configuración de empresas no encontrada'}

            company_folders = config['company_folders']
            if not company_folders:
                return {'success': False, 'error': 'No hay carpetas empresariales configuradas'}

            combustible_config = config.get('combustible_exclusions', {})
            emitter_names = []
            if isinstance(combustible_config, dict):
                emitter_names = combustible_config.get('emitter_names', [])
            elif isinstance(combustible_config, list):
                emitter_names = combustible_config

            self.combustible_exclusion_emitters = {
                self._normalize_emisor_name(name)
                for name in emitter_names
                if isinstance(name, str) and name.strip()
            }

            if self.combustible_exclusion_emitters:
                self.log_message(
                    f"⛽️ Exclusiones de combustible activas: {len(self.combustible_exclusion_emitters)} emisores",
                    "info"
                )

            if not self._validate_folders(input_folder, output_folder, company_folders):
                return {'success': False, 'error': 'Carpetas inválidas'}

            if self.stop_event.is_set():
                return {'success': False, 'error': 'Procesamiento cancelado'}

            excel_files = self._find_excel_files(input_folder)
            if not excel_files:
                return {'success': False, 'error': 'No se encontraron archivos Excel'}

            # Usar rutas dinámicas para construir índices XML
            xml_data_by_company = self._build_xml_indices_by_company_dynamic(company_folders)
            if not xml_data_by_company:
                return {'success': False, 'error': 'No hay archivos XML válidos en ninguna empresa'}

            if self.stop_event.is_set():
                return {'success': False, 'error': 'Procesamiento cancelado'}

            results = []
            manual_review_limit = config.get('manual_review_limit', 3)
            delete_originals = config.get('delete_originals', True)
            commercial_activities = config.get('commercial_activities', {})

            for i, excel_file in enumerate(excel_files, 1):
                if self.stop_event.is_set():
                    self.log_message("⏹️ Procesamiento cancelado por el usuario", "warning")
                    break

                self.log_message(f"📊 Procesando Excel {i}/{len(excel_files)}: {os.path.basename(excel_file)}", "info")

                excel_results = self._process_excel_against_all_companies(
                    excel_file, xml_data_by_company, output_folder, manual_review_limit, commercial_activities
                )

                results.extend(excel_results)

                if delete_originals and excel_results:
                    try:
                        os.remove(excel_file)
                        self.log_message(f"🗑️ Archivo original eliminado: {os.path.basename(excel_file)}", "info")
                    except Exception as e:
                        self.log_message(f"⚠️ No se pudo eliminar original: {str(e)}", "warning")

            self.stats['processing_time'] = time.time() - start_time
            self.stats['files_created'] = len(results)

            # Calcular tasa de extracción de placas
            if self.stats['otro_texto_processed'] > 0:
                self.stats['placa_extraction_rate'] = (self.stats['placas_extracted'] / self.stats[
                    'otro_texto_processed']) * 100

            if not self.stop_event.is_set():
                self._log_processing_summary()

            return {
                'success': len(results) > 0,
                'processed_files': results,
                'stats': self.stats.copy(),
                'stopped_by_user': self.stop_event.is_set()
            }

        except Exception as e:
            self.log_message(f"❌ Error crítico en procesamiento: {str(e)}", "error")
            return {'success': False, 'error': str(e)}

    def stop_processing(self):
        """Detiene el procesamiento en curso."""
        self.stop_event.set()
        self.log_message("⏹️ Señal de parada enviada al procesador", "warning")

    def _reset_stats(self):
        """Resetea las estadísticas para un nuevo procesamiento."""
        self.stats = {
            k: {} if k == 'company_details' else [] if k in ['companies_no_matches', 'excluded_by_date_details',
                                                             'companies_folders_skipped',
                                                             'excel_excluded_by_date_details'] else 0
            for k in self.stats}
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year

    def _parse_xml_date(self, fecha_emision_text: str) -> Optional[datetime]:
        """
        Parsea la fecha de emisión del XML desde formato ISO.

        Args:
            fecha_emision_text (str): Fecha en formato '2025-07-01T10:19:14-06:00'

        Returns:
            datetime: Objeto datetime o None si no se puede parsear
        """
        try:
            # Formato típico: 2025-07-01T10:19:14-06:00
            # Extraer solo la parte de fecha y hora, ignorar timezone
            if 'T' in fecha_emision_text:
                fecha_part = fecha_emision_text.split('T')[0]
                return datetime.strptime(fecha_part, '%Y-%m-%d')
            else:
                # Si solo tiene fecha
                return datetime.strptime(fecha_emision_text[:10], '%Y-%m-%d')
        except (ValueError, IndexError) as e:
            return None

    def _is_current_month(self, fecha_emision: datetime) -> bool:
        """
        Verifica si una fecha pertenece al mes y año actual.

        Args:
            fecha_emision (datetime): Fecha a verificar

        Returns:
            bool: True si es del mes actual
        """
        return (fecha_emision.month == self.current_month and
                fecha_emision.year == self.current_year)

    def _validate_folders(self, input_folder: str, output_folder: str, company_folders: Dict) -> bool:
        """Valida que las carpetas existan y sean accesibles (solo carpetas base para rutas dinámicas)."""
        try:
            if not os.path.exists(input_folder):
                self.log_message(f"❌ Carpeta de entrada no existe: {input_folder}", "error")
                return False

            # Para rutas dinámicas, validamos solo las carpetas base
            for company_key, base_xml_folder in company_folders.items():
                if not os.path.exists(base_xml_folder):
                    self.log_message(f"❌ Carpeta BASE no existe para {company_key}: {base_xml_folder}", "error")
                    return False
                if not os.access(base_xml_folder, os.R_OK):
                    self.log_message(f"❌ Carpeta BASE no accesible para {company_key}: {base_xml_folder}", "error")
                    return False

            os.makedirs(output_folder, exist_ok=True)
            return True

        except Exception as e:
            self.log_message(f"❌ Error validando carpetas: {str(e)}", "error")
            return False

    def _find_excel_files(self, input_folder: str) -> List[str]:
        """Busca archivos Excel en la carpeta de entrada."""
        try:
            excel_files = glob.glob(os.path.join(input_folder, "cargador*.xlsx"))

            if excel_files:
                self.log_message(f"📊 Encontrados {len(excel_files)} archivos Excel para procesar", "success")
                for i, file in enumerate(excel_files, 1):
                    self.log_message(f"  {i}. {os.path.basename(file)}", "info")
            else:
                self.log_message("📭 No se encontraron archivos Excel con patrón 'cargador*.xlsx'", "warning")

            return excel_files

        except Exception as e:
            self.log_message(f"❌ Error buscando archivos Excel: {str(e)}", "error")
            return []

    def _build_xml_indices_by_company_dynamic(self, company_folders: Dict) -> Dict[str, Dict[str, List[str]]]:
        """Construye índices XML separados para cada empresa usando rutas dinámicas automáticas con manejo graceful de carpetas faltantes y soporte para extracción de placas desde OtroTexto."""

        # Importar ConfigManager para usar las funciones de rutas dinámicas
        from contaflow.config.config_manager import ConfigManager
        config_manager = ConfigManager()

        current_date = datetime.now()
        self.log_message(
            f"📂 Construyendo índices XML por empresa con rutas dinámicas (mes actual: {current_date.month}/{current_date.year}) con soporte para Correos y extracción de placas desde OtroTexto...",
            "info")

        # Obtener información de rutas dinámicas
        dynamic_paths_info = config_manager.get_all_dynamic_xml_paths(company_folders)

        xml_data_by_company = {}
        self.stats['companies_processed'] = len(company_folders)

        for company_key, path_info in dynamic_paths_info.items():
            if self.stop_event.is_set():
                break

            company_name = self._get_company_display_name(company_key)

            # Inicializar estadísticas de empresa
            self.stats['company_details'][company_key] = {
                'name': company_name, 'xml_count': 0, 'xml_valid': 0,
                'xml_current_month': 0, 'xml_excluded_by_date': 0,
                'xml_errors': 0, 'matches': 0, 'manual_reviews': 0,
                'correos_processed': 0, 'placas_extracted': 0,
                'combustible_exclusions': 0,
                'folder_exists': path_info['exists'], 'dynamic_path': path_info['dynamic_path'],
                'excel_rows_processed': 0, 'excel_rows_excluded': 0
            }

            # Verificar si la carpeta del mes actual existe
            if not path_info['exists']:
                # Carpeta del mes actual no existe - SALTAR sin error
                self.stats['companies_folders_missing'] += 1
                self.stats['companies_folders_skipped'].append(company_name)
                self.log_message(
                    f"📅 {company_name}: Sin carpeta {current_date.month:02d}/{current_date.year} - Se omite del procesamiento",
                    "info")
                continue

            # Carpeta del mes actual existe - PROCESAR
            self.stats['companies_folders_found'] += 1
            self.log_message(
                f"🏢 Procesando empresa: {company_name} (carpeta {current_date.month:02d}/{current_date.year})", "info")

            xml_data = self._build_xml_index_for_company(path_info['dynamic_path'], company_key)

            if xml_data:
                xml_data_by_company[company_key] = xml_data
                self.stats['companies_with_matches'] += 1
                current_month_count = self.stats['company_details'][company_key]['xml_current_month']
                correos_count = self.stats['company_details'][company_key]['correos_processed']
                placas_count = self.stats['company_details'][company_key]['placas_extracted']
                self.log_message(
                    f"✅ {company_name}: {len(xml_data)} números únicos indexados (mes actual, {correos_count} de Correos, {placas_count} placas extraídas)",
                    "success")
            else:
                self.stats['companies_without_matches'] += 1
                self.stats['companies_no_matches'].append(company_name)
                self.log_message(f"⚠️ {company_name}: Sin archivos XML válidos del mes actual", "warning")

        # Log resumen de rutas dinámicas
        excluded_total = self.stats['total_xml_excluded_by_date']
        current_total = self.stats['total_xml_current_month']
        correos_total = self.stats['correos_pdfs_processed']
        placas_total = self.stats['placas_extracted']
        folders_found = self.stats['companies_folders_found']
        folders_missing = self.stats['companies_folders_missing']

        self.log_message(
            f"📊 Rutas dinámicas procesadas: {folders_found} carpetas encontradas, {folders_missing} carpetas faltantes",
            "info")
        self.log_message(
            f"📊 Filtrado XML por fecha completado: {current_total} XMLs del mes actual, {excluded_total} excluidos, {correos_total} de Correos procesados, {placas_total} placas extraídas",
            "info")

        if self.stats['companies_folders_skipped']:
            self.log_message(
                f"📅 Empresas omitidas (sin carpeta {current_date.month:02d}/{current_date.year}): {', '.join(self.stats['companies_folders_skipped'])}",
                "info")

        return xml_data_by_company

    def _build_xml_index_for_company(self, xml_folder: str, company_key: str) -> Dict[str, List[str]]:
        """Construye índice XML para una empresa específica con filtrado por fecha, soporte para Correos, y extracción de placas desde OtroTexto."""
        xml_files_info = self._find_xml_files_recursive(xml_folder)
        company_stats = self.stats['company_details'][company_key]
        company_stats['xml_count'] = len(xml_files_info)
        self.stats['total_xml_count'] += len(xml_files_info)

        if not xml_files_info:
            return {}

        xml_data = {}

        for i, (xml_file, rel_path, depth) in enumerate(xml_files_info, 1):
            if self.stop_event.is_set():
                break

            try:
                if i % 200 == 0 or i <= 10 or i > len(xml_files_info) - 5:
                    progress = (i / len(xml_files_info)) * 100
                    self.log_message(f"   [{progress:5.1f}%] {rel_path}", "info")

                tree = ET.parse(xml_file)
                root = tree.getroot()

                numero = None
                fecha_emision = None

                # Extraer NumeroConsecutivo y FechaEmision
                for elem in root.iter():
                    if 'NumeroConsecutivo' in elem.tag and elem.text:
                        numero = elem.text.strip()
                    elif 'FechaEmision' in elem.tag and elem.text:
                        fecha_emision = elem.text.strip()

                if numero and fecha_emision:
                    # FILTRADO POR FECHA - Proceso temprano crítico
                    parsed_date = self._parse_xml_date(fecha_emision)

                    if parsed_date is None:
                        # Error parseando fecha, contar como error
                        company_stats['xml_errors'] += 1
                        continue

                    if not self._is_current_month(parsed_date):
                        # XML fuera del mes actual - EXCLUIR del procesamiento
                        company_stats['xml_excluded_by_date'] += 1
                        self.stats['total_xml_excluded_by_date'] += 1

                        # Registrar para el reporte PDF
                        self.stats['excluded_by_date_details'].append({
                            'company': self._get_company_display_name(company_key),
                            'numero_consecutivo': numero,
                            'fecha_emision': fecha_emision,
                            'fecha_parsed': parsed_date.strftime('%Y-%m-%d')
                        })
                        continue  # SALTAR este XML completamente

                    # XML del mes actual - INCLUIR en el análisis
                    # VERIFICAR SI ES FACTURA DE CORREOS (tiene prioridad sobre extracción de placas)
                    is_correos = self._is_correos_xml(xml_file)

                    if is_correos:
                        # Procesar como factura de Correos usando PDF
                        success, correos_data = self._process_correos_xml(xml_file, numero)
                        if success and correos_data:
                            xml_data[numero] = [correos_data]  # Usar dato del PDF
                            company_stats['correos_processed'] += 1
                        else:
                            # Si falla el procesamiento de PDF, usar detalles normales como fallback
                            detalles = [elem.text.strip() for elem in root.iter()
                                        if 'Detalle' in elem.tag and elem.text and elem.text.strip()]
                            xml_data[numero] = detalles
                    else:
                        if not self._apply_combustible_exclusion_if_needed(root, numero, xml_data, company_stats):
                            # Intentar extraer placa desde OtroTexto
                            placa_info = self._extract_otro_texto_info(root)

                            if placa_info:
                                # Se extrajo placa exitosamente desde OtroTexto
                                xml_data[numero] = [placa_info]
                                company_stats['placas_extracted'] += 1
                            else:
                                # No se pudo extraer placa o no hay OtroTexto - usar Detalle como fallback
                                detalles = [elem.text.strip() for elem in root.iter()
                                            if 'Detalle' in elem.tag and elem.text and elem.text.strip()]
                                xml_data[numero] = detalles
                                self.stats['fallback_to_detalle'] += 1

                    company_stats['xml_valid'] += 1
                    company_stats['xml_current_month'] += 1
                    self.stats['total_xml_valid'] += 1
                    self.stats['total_xml_current_month'] += 1

                elif numero and not fecha_emision:
                    # Tiene numero pero no fecha - incluir con advertencia
                    # Verificar si es de Correos
                    is_correos = self._is_correos_xml(xml_file)

                    if is_correos:
                        # Procesar como factura de Correos usando PDF
                        success, correos_data = self._process_correos_xml(xml_file, numero)
                        if success and correos_data:
                            xml_data[numero] = [correos_data]  # Usar dato del PDF
                            company_stats['correos_processed'] += 1
                        else:
                            # Fallback a detalles normales
                            detalles = [elem.text.strip() for elem in root.iter()
                                        if 'Detalle' in elem.tag and elem.text and elem.text.strip()]
                            xml_data[numero] = detalles
                    else:
                        if not self._apply_combustible_exclusion_if_needed(root, numero, xml_data, company_stats):
                            # Intentar extraer placa desde OtroTexto
                            placa_info = self._extract_otro_texto_info(root)

                            if placa_info:
                                # Se extrajo placa exitosamente desde OtroTexto
                                xml_data[numero] = [placa_info]
                                company_stats['placas_extracted'] += 1
                            else:
                                # No se pudo extraer placa o no hay OtroTexto - usar Detalle como fallback
                                detalles = [elem.text.strip() for elem in root.iter()
                                            if 'Detalle' in elem.tag and elem.text and elem.text.strip()]
                                xml_data[numero] = detalles
                                self.stats['fallback_to_detalle'] += 1

                    company_stats['xml_valid'] += 1
                    company_stats['xml_current_month'] += 1
                    self.stats['total_xml_valid'] += 1
                    self.stats['total_xml_current_month'] += 1

            except (ET.ParseError, Exception) as e:
                company_stats['xml_errors'] += 1
                if company_stats['xml_errors'] <= 3:
                    self.log_message(f"⚠️ Error XML: {rel_path} - {str(e)[:50]}", "warning")

        return xml_data

    def _find_xml_files_recursive(self, root_folder: str) -> List[Tuple[str, str, int]]:
        """Encuentra todos los archivos XML de manera recursiva."""
        xml_files = []

        try:
            for root, dirs, files in os.walk(root_folder):
                if self.stop_event.is_set():
                    break

                relative_path = os.path.relpath(root, root_folder)
                depth = 0 if relative_path == '.' else len(relative_path.split(os.sep))

                xml_files.extend([
                    (os.path.join(root, f), os.path.relpath(os.path.join(root, f), root_folder), depth)
                    for f in files if f.lower().endswith('.xml')
                ])

        except (PermissionError, Exception) as e:
            self.log_message(f"⚠️ Error explorando directorios: {str(e)[:50]}", "warning")

        return xml_files

    def _process_excel_against_all_companies(self, excel_file: str, xml_data_by_company: Dict,
                                             output_folder: str, manual_review_limit: int = 3,
                                             commercial_activities: Dict = None) -> List[Dict]:
        """Procesa un archivo Excel contra todas las empresas con filtrado por fecha en las filas del Excel."""
        filename = os.path.basename(excel_file)
        results = []

        # Asegurar que commercial_activities no sea None
        if commercial_activities is None:
            commercial_activities = {}

        try:
            if self.stop_event.is_set():
                return results

            wb = load_workbook(excel_file, read_only=True)
            sheet = wb.active
            excel_rows = list(sheet.iter_rows(min_row=2, values_only=True))
            wb.close()

            # Filtrar filas de Excel por fecha ANTES de procesarlas contra las empresas
            filtered_excel_rows = self._filter_excel_rows_by_date(excel_rows, filename)

            for company_key, xml_data in xml_data_by_company.items():
                if self.stop_event.is_set():
                    break

                company_name = self._get_company_display_name(company_key)
                self.log_message(f"  🏢 Procesando contra: {company_name}", "info")

                # Obtener actividad comercial para esta empresa (puede estar vacía)
                company_activity = commercial_activities.get(company_key, "")

                result = self._process_excel_for_company(
                    filtered_excel_rows, xml_data, company_key, output_folder,
                    filename, manual_review_limit, company_activity
                )

                if result:
                    results.append(result)
                    self.log_message(f"  ✅ {company_name}: Archivo generado con {result['matches']} matches", "success")
                else:
                    if company_name not in self.stats['companies_no_matches']:
                        self.stats['companies_no_matches'].append(company_name)
                    self.log_message(f"  ❌ {company_name}: Sin matches encontrados", "warning")

            if results:
                self.stats['excel_processed'] += 1

            return results

        except Exception as e:
            self.log_message(f"❌ Error procesando Excel {filename}: {str(e)}", "error")
            return results

    def _filter_excel_rows_by_date(self, excel_rows: List, filename: str) -> List:
        """
        Filtra las filas del Excel por la fecha del documento, manteniendo solo las del mes actual.

        Args:
            excel_rows (List): Lista de filas del Excel
            filename (str): Nombre del archivo para logging

        Returns:
            List: Filas filtradas que pertenecen al mes actual
        """
        filtered_rows = []
        total_rows = len(excel_rows)
        excluded_count = 0

        self.log_message(
            f"📅 Filtrando filas del Excel por fecha (mes actual: {self.current_month}/{self.current_year})", "info")

        for i, row in enumerate(excel_rows):
            if not row or len(row) <= self.fecha_documento_column_index:
                # Fila vacía o sin suficientes columnas
                continue

            fecha_documento_value = row[self.fecha_documento_column_index]

            if fecha_documento_value:
                parsed_date = self._parse_excel_date(fecha_documento_value)

                if parsed_date is None:
                    # No se pudo parsear la fecha - incluir con advertencia
                    self.log_message(f"⚠️ Fecha no válida en fila {i + 1}: '{fecha_documento_value}'", "warning")
                    filtered_rows.append(row)
                    continue

                if not self._is_excel_row_current_month(parsed_date):
                    # Fecha fuera del mes actual - EXCLUIR
                    excluded_count += 1
                    self.stats['excel_rows_excluded_by_date'] += 1

                    # Registrar para estadísticas detalladas
                    numero_consecutivo = str(row[1]).strip() if len(row) > 1 and row[1] else "N/A"
                    self.stats['excel_excluded_by_date_details'].append({
                        'filename': filename,
                        'numero_consecutivo': numero_consecutivo,
                        'fecha_documento': str(fecha_documento_value),
                        'fecha_parsed': parsed_date.strftime('%d-%m-%Y')
                    })
                    continue

                # Fecha del mes actual - INCLUIR
                filtered_rows.append(row)
            else:
                # Sin fecha - incluir con advertencia
                self.log_message(f"⚠️ Fila {i + 1} sin fecha en columna 'Fecha Documento'", "warning")
                filtered_rows.append(row)

        # Actualizar estadísticas globales
        self.stats['excel_rows_total'] += total_rows
        self.stats['excel_rows_current_month'] += len(filtered_rows)

        # Log resultado del filtrado
        if excluded_count > 0:
            self.log_message(
                f"📊 Excel filtrado: {len(filtered_rows)} filas del mes actual, {excluded_count} excluidas por fecha",
                "info"
            )
        else:
            self.log_message(f"📊 Excel procesado: {len(filtered_rows)} filas válidas", "info")

        return filtered_rows

    def _process_excel_for_company(self, excel_rows: List, xml_data: Dict, company_key: str,
                                   output_folder: str, original_filename: str,
                                   manual_review_limit: int = 3, company_activity: str = "") -> Optional[Dict]:
        """Procesa un Excel para una empresa específica incluyendo actividad comercial y campo Paquete."""
        try:
            if self.stop_event.is_set():
                return None

            new_wb = Workbook()
            new_sheet = new_wb.active
            self._setup_excel_headers(new_sheet)

            row_out = 2
            matches = 0
            manual_reviews = 0
            company_stats = self.stats['company_details'][company_key]

            for row in excel_rows:
                if self.stop_event.is_set() or not row or not row[1]:
                    continue

                numero = str(row[1]).strip()

                # Buscar el número en los datos XML (sin verificar duplicados)
                if numero in xml_data:
                    detalles = xml_data[numero]

                    if len(detalles) > manual_review_limit:
                        self._write_excel_row_manual_review(new_sheet, row_out, row, company_activity)
                        manual_reviews += 1
                        company_stats['manual_reviews'] += 1
                        self.stats['total_manual_reviews'] += 1
                    else:
                        detalle_text = " | ".join(detalles) if detalles else ""
                        self._write_excel_row_with_detail(new_sheet, row_out, row, detalle_text, company_activity)

                    row_out += 1
                    matches += 1

            if matches > 0 and not self.stop_event.is_set():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_name = original_filename.split('.')[0]
                company_file_name = self.company_file_names.get(company_key, company_key)
                output_file = os.path.join(
                    output_folder,
                    f"{base_name}_procesado_{company_file_name}_{timestamp}.xlsx"
                )

                self._adjust_excel_column_widths(new_sheet)
                new_wb.save(output_file)
                new_wb.close()

                company_stats['matches'] += matches
                company_stats['excel_rows_processed'] += len(excel_rows)
                self.stats['total_matches'] += matches

                return {
                    'source': original_filename,
                    'output': output_file,
                    'company_key': company_key,
                    'company_name': self._get_company_display_name(company_key),
                    'matches': matches,
                    'manual_reviews': manual_reviews,
                    'filename': os.path.basename(output_file),
                    'commercial_activity': company_activity
                }
            else:
                new_wb.close()
                return None

        except Exception as e:
            self.log_message(f"❌ Error procesando empresa {company_key}: {str(e)}", "error")
            return None

    def _get_company_display_name(self, company_key: str) -> str:
        """Obtiene el nombre de display de una empresa."""
        return {
            'nargallo': 'Nargallo del Este S.A.',
            'ventas_fruno': 'Ventas Fruno, S.A.',
            'creme_caramel': 'Creme Caramel',
            'su_laka': 'Su Laka'
        }.get(company_key, company_key)

    def _setup_excel_headers(self, sheet):
        """Configura los headers de Excel con estilo incluyendo Paquete y Actividad Comercial."""
        for col, header in enumerate(self.excel_columns, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_excel_row_with_detail(self, sheet, row_num: int, row_data: tuple, detail: str, company_activity: str):
        """Escribe una fila de Excel con el detalle del XML, campo Paquete y actividad comercial."""
        # Procesar las primeras 17 columnas (incluyendo Paquete)
        for col, value in enumerate(row_data[:17], 1):
            if col == 6:  # Columna Aplicación
                sheet.cell(row=row_num, column=col, value=detail)
            else:
                sheet.cell(row=row_num, column=col, value=value)

        # Columna 18: Actividad Comercial (la última)
        sheet.cell(row=row_num, column=18, value=company_activity)

    def _write_excel_row_manual_review(self, sheet, row_num: int, row_data: tuple, company_activity: str):
        """Escribe una fila de Excel marcada para revisión manual con Paquete y actividad comercial."""
        # Procesar las primeras 17 columnas (incluyendo Paquete)
        for col, value in enumerate(row_data[:17], 1):
            cell = sheet.cell(row=row_num, column=col)
            if col == 6:  # Columna Aplicación
                cell.value = "Revision Manual"
                cell.font = self.red_font  # Font ya tiene color blanco
                cell.fill = self.red_fill
            else:
                cell.value = value

        # Columna 18: Actividad Comercial (la última)
        sheet.cell(row=row_num, column=18, value=company_activity)

    def _adjust_excel_column_widths(self, sheet):
        """Ajusta automáticamente el ancho de las columnas incluyendo Paquete y Actividad Comercial."""
        for column in sheet.columns:
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

    def _log_processing_summary(self):
        """Registra un resumen detallado del procesamiento con información de rutas dinámicas, filtrado por fecha en XMLs y Excel, estadísticas de Correos, y nueva funcionalidad de extracción de placas desde OtroTexto."""
        self.log_message("=" * 60, "info")
        self.log_message(
            "📋 RESUMEN DE PROCESAMIENTO EMPRESARIAL CON RUTAS DINÁMICAS, FILTRADO POR FECHA, SOPORTE CORREOS Y EXTRACCIÓN DE PLACAS",
            "info")
        self.log_message("=" * 60, "info")

        summaries = [
            ("🏢 EMPRESAS:", [
                (f"📊 Empresas configuradas: {self.stats['companies_processed']}",),
                (f"📂 Carpetas mes actual encontradas: {self.stats['companies_folders_found']}",),
                (f"📅 Carpetas mes actual faltantes: {self.stats['companies_folders_missing']}",),
                (f"✅ Empresas con matches: {self.stats['companies_with_matches']}",),
                (f"❌ Empresas sin matches: {self.stats['companies_without_matches']}",)
            ]),
            ("📅 FILTRADO XML POR FECHA:", [
                (f"📄 XML encontrados: {self.stats['total_xml_count']}",),
                (f"✅ XML válidos: {self.stats['total_xml_valid']}",),
                (f"📅 XML del mes actual: {self.stats['total_xml_current_month']}",),
                (f"🚫 XML excluidos por fecha: {self.stats['total_xml_excluded_by_date']}",)
            ]),
            ("📊 FILTRADO EXCEL POR FECHA:", [
                (f"📄 Filas Excel encontradas: {self.stats['excel_rows_total']}",),
                (f"📅 Filas Excel del mes actual: {self.stats['excel_rows_current_month']}",),
                (f"🚫 Filas Excel excluidas por fecha: {self.stats['excel_rows_excluded_by_date']}",)
            ]),
            ("📮 CORREOS DE COSTA RICA SA:", [
                (f"📄 PDFs de Correos procesados: {self.stats['correos_pdfs_processed']}",),
                (f"❌ PDFs de Correos fallidos: {self.stats['correos_pdfs_failed']}",),
                (f"✅ Matches de Correos: {self.stats['correos_matches']}",)
            ]),
            ("🚗 EXTRACCIÓN DE PLACAS DESDE OTROTEXTO:", [
                (f"📄 Campos OtroTexto procesados: {self.stats['otro_texto_processed']}",),
                (f"🚗 Placas extraídas exitosamente: {self.stats['placas_extracted']}",),
                (f"❌ Extracciones fallidas: {self.stats['placas_failed']}",),
                (f"📋 Fallback a Detalle: {self.stats['fallback_to_detalle']}",),
                (f"⛽️ Exclusiones de combustible aplicadas: {self.stats['combustible_exclusions_applied']}",),
                (f"📈 Tasa de extracción: {self.stats['placa_extraction_rate']:.1f}%",)
            ]),
            ("⚙️ PROCESAMIENTO:", [
                (f"📊 Excel procesados: {self.stats['excel_processed']}",),
                (f"✅ Matches exitosos: {self.stats['total_matches']}",),
                (f"🔴 Revisiones manuales: {self.stats['total_manual_reviews']}",),
                (f"📁 Archivos creados: {self.stats['files_created']}",)
            ])
        ]

        for title, items in summaries:
            self.log_message(title, "info")
            for item in items:
                self.log_message(f"   {item[0]}", "info")

        self.log_message("🏢 DETALLES POR EMPRESA:", "info")
        for company_key, details in self.stats['company_details'].items():
            status = "✅" if details['matches'] > 0 else "❌"
            folder_status = "📂" if details['folder_exists'] else "📅"
            excluded = details['xml_excluded_by_date']
            current = details['xml_current_month']
            correos = details.get('correos_processed', 0)
            placas = details.get('placas_extracted', 0)
            exclusions = details.get('combustible_exclusions', 0)
            excel_rows = details.get('excel_rows_processed', 0)
            extras = []
            if correos > 0:
                extras.append(f"{correos} Correos")
            if placas > 0:
                extras.append(f"{placas} placas")
            if exclusions > 0:
                extras.append(f"{exclusions} excl. combustible")
            extras_text = f", {', '.join(extras)}" if extras else ""
            self.log_message(
                f"   {status} {folder_status} {details['name']}: {details['matches']} matches ({current} XMLs analizados, {excluded} excluidos{extras_text})",
                "info")

        if self.stats['companies_folders_skipped']:
            current_date = datetime.now()
            self.log_message("📅 EMPRESAS OMITIDAS (sin carpeta del mes actual):", "info")
            for company in self.stats['companies_folders_skipped']:
                self.log_message(f"   • {company} (falta carpeta {current_date.month:02d}/{current_date.year})", "info")

        if self.stats['companies_no_matches']:
            self.log_message("⚠️ EMPRESAS SIN MATCHES:", "warning")
            for company in self.stats['companies_no_matches']:
                self.log_message(f"   • {company}", "warning")

        if self.stats['excluded_by_date_details']:
            self.log_message("🚫 XMLS EXCLUIDOS POR FECHA (primeros 10):", "info")
            for i, excluded in enumerate(self.stats['excluded_by_date_details'][:10], 1):
                self.log_message(
                    f"   {i}. {excluded['numero_consecutivo']} ({excluded['company']}) - {excluded['fecha_parsed']}",
                    "info")

            if len(self.stats['excluded_by_date_details']) > 10:
                remaining = len(self.stats['excluded_by_date_details']) - 10
                self.log_message(f"   ... y {remaining} más", "info")

        if self.stats['excel_excluded_by_date_details']:
            self.log_message("🚫 FILAS EXCEL EXCLUIDAS POR FECHA (primeros 10):", "info")
            for i, excluded in enumerate(self.stats['excel_excluded_by_date_details'][:10], 1):
                self.log_message(
                    f"   {i}. {excluded['numero_consecutivo']} ({excluded['filename']}) - {excluded['fecha_parsed']}",
                    "info")

            if len(self.stats['excel_excluded_by_date_details']) > 10:
                remaining = len(self.stats['excel_excluded_by_date_details']) - 10
                self.log_message(f"   ... y {remaining} más", "info")

        time_str = f"{self.stats['processing_time'] / 60:.1f} minutos" if self.stats[
                                                                              'processing_time'] > 60 else f"{self.stats['processing_time']:.1f} segundos"
        self.log_message(f"   ⏱️ Tiempo total: {time_str}", "info")
        self.log_message(
            "📦 ACTUALIZACIONES: Rutas dinámicas automáticas, filtrado por fecha en XMLs y Excel, campo 'Paquete', soporte para PDFs de Correos, y extracción de placas desde OtroTexto",
            "success")
        self.log_message("=" * 60, "info")

    def log_message(self, message: str, msg_type: str = "info"):
        """Envía un mensaje al log de la interfaz de forma segura."""
        try:
            if self.automation_tab and hasattr(self.automation_tab, 'add_log_message'):
                self.automation_tab.add_log_message(message, msg_type)
            else:
                print(f"[{msg_type.upper()}] {message}")
        except Exception:
            print(f"[{msg_type.upper()}] {message}")

    def get_processing_stats(self) -> Dict:
        """Obtiene las estadísticas actuales del procesamiento."""
        return self.stats.copy()

    def get_companies_without_matches(self) -> List[str]:
        """Obtiene la lista de empresas que no tuvieron matches."""
        return self.stats['companies_no_matches'].copy()

    def get_excluded_by_date_details(self) -> List[Dict]:
        """Obtiene la lista detallada de XMLs excluidos por fecha."""
        return self.stats['excluded_by_date_details'].copy()

    def get_excel_excluded_by_date_details(self) -> List[Dict]:
        """Obtiene la lista detallada de filas de Excel excluidas por fecha."""
        return self.stats['excel_excluded_by_date_details'].copy()

    def get_correos_stats(self) -> Dict:
        """Obtiene estadísticas específicas del procesamiento de Correos."""
        return {
            'correos_pdfs_processed': self.stats['correos_pdfs_processed'],
            'correos_pdfs_failed': self.stats['correos_pdfs_failed'],
            'correos_matches': self.stats['correos_matches']
        }

    def get_dynamic_paths_stats(self) -> Dict:
        """Obtiene estadísticas específicas del procesamiento con rutas dinámicas."""
        return {
            'companies_folders_found': self.stats['companies_folders_found'],
            'companies_folders_missing': self.stats['companies_folders_missing'],
            'companies_folders_skipped': self.stats['companies_folders_skipped'].copy()
        }

    def get_excel_filtering_stats(self) -> Dict:
        """Obtiene estadísticas específicas del filtrado de Excel por fecha."""
        return {
            'excel_rows_total': self.stats['excel_rows_total'],
            'excel_rows_current_month': self.stats['excel_rows_current_month'],
            'excel_rows_excluded_by_date': self.stats['excel_rows_excluded_by_date'],
            'excel_excluded_by_date_details': self.stats['excel_excluded_by_date_details'].copy()
        }

    def get_placa_extraction_stats(self) -> Dict:
        """Obtiene estadísticas específicas de la extracción de placas desde OtroTexto."""
        return {
            'otro_texto_processed': self.stats['otro_texto_processed'],
            'placas_extracted': self.stats['placas_extracted'],
            'placas_failed': self.stats['placas_failed'],
            'fallback_to_detalle': self.stats['fallback_to_detalle'],
            'placa_extraction_rate': self.stats['placa_extraction_rate']
        }

    def get_detailed_report(self) -> str:
        """Genera un reporte detallado del procesamiento empresarial con información de rutas dinámicas, filtrado por fecha en XMLs y Excel, Correos, y extracción de placas desde OtroTexto."""
        time_str = f"{self.stats['processing_time'] / 60:.1f} min" if self.stats[
                                                                          'processing_time'] > 60 else f"{self.stats['processing_time']:.1f} seg"

        current_date = datetime.now()

        report = f"""📊 REPORTE DETALLADO DE PROCESAMIENTO EMPRESARIAL CON RUTAS DINÁMICAS, FILTRADO POR FECHA Y EXTRACCIÓN DE PLACAS

🏢 EMPRESAS:
• Configuradas: {self.stats['companies_processed']}
• Carpetas {current_date.month:02d}/{current_date.year} encontradas: {self.stats['companies_folders_found']}
• Carpetas {current_date.month:02d}/{current_date.year} faltantes: {self.stats['companies_folders_missing']}
• Con matches: {self.stats['companies_with_matches']}
• Sin matches: {self.stats['companies_without_matches']}

📅 FILTRADO XML POR FECHA (mes actual: {self.current_month}/{self.current_year}):
• XML encontrados: {self.stats['total_xml_count']}
• XML del mes actual: {self.stats['total_xml_current_month']}
• XML excluidos por fecha: {self.stats['total_xml_excluded_by_date']}

📊 FILTRADO EXCEL POR FECHA:
• Filas Excel encontradas: {self.stats['excel_rows_total']}
• Filas Excel del mes actual: {self.stats['excel_rows_current_month']}
• Filas Excel excluidas por fecha: {self.stats['excel_rows_excluded_by_date']}

📮 CORREOS DE COSTA RICA SA:
• PDFs procesados: {self.stats['correos_pdfs_processed']}
• PDFs fallidos: {self.stats['correos_pdfs_failed']}
• Matches de Correos: {self.stats['correos_matches']}

🚗 EXTRACCIÓN DE PLACAS DESDE OTROTEXTO:
• Campos OtroTexto procesados: {self.stats['otro_texto_processed']}
• Placas extraídas exitosamente: {self.stats['placas_extracted']}
• Extracciones fallidas: {self.stats['placas_failed']}
• Fallback a Detalle: {self.stats['fallback_to_detalle']}
• Tasa de extracción: {self.stats['placa_extraction_rate']:.1f}%

⚙️ PROCESAMIENTO:
• XML válidos analizados: {self.stats['total_xml_valid']}
• Excel procesados: {self.stats['excel_processed']}

🎯 RESULTADOS:
• Matches exitosos: {self.stats['total_matches']}
• Revisiones manuales: {self.stats['total_manual_reviews']}
• Archivos creados: {self.stats['files_created']}

📈 RENDIMIENTO:
• Tiempo total: {time_str}
• Estado: {'✅ Completado' if not self.stop_event.is_set() else '⏹️ Cancelado'}

📦 ACTUALIZACIONES:
• Rutas dinámicas automáticas basadas en año/mes actual
• Manejo graceful de carpetas faltantes sin fallar
• Campo 'Paquete' incluido en archivos procesados
• Procesamiento especial para facturas de Correos de Costa Rica SA
• Extracción automática de datos desde PDFs de Correos
• Filtrado por fecha tanto en XMLs como en filas de Excel
• NUEVA: Extracción de códigos de placas desde campo <OtroTexto>
• NUEVA: Formato "Combustible / Placa: [CÓDIGO]" para placas
• NUEVA: Fallback automático a <Detalle> si no se encuentra placa
• Reporte de elementos excluidos por fecha
"""

        if self.stats['company_details']:
            report += "\n🏢 DETALLES POR EMPRESA:\n"
            for details in self.stats['company_details'].values():
                status = "✅" if details['matches'] > 0 else "❌"
                folder_status = "📂" if details['folder_exists'] else "📅"
                excluded = details['xml_excluded_by_date']
                current = details['xml_current_month']
                correos = details.get('correos_processed', 0)
                placas = details.get('placas_extracted', 0)
                extras = []
                if correos > 0:
                    extras.append(f"{correos} Correos")
                if placas > 0:
                    extras.append(f"{placas} placas")
                extras_text = f", {', '.join(extras)}" if extras else ""
                report += f"• {status} {folder_status} {details['name']}: {details['matches']} matches ({current} XMLs, {excluded} excluidos{extras_text})\n"

        if self.stats['companies_folders_skipped']:
            report += f"\n📅 EMPRESAS OMITIDAS (sin carpeta {current_date.month:02d}/{current_date.year}):\n"
            report += "\n".join(f"• {company}" for company in self.stats['companies_folders_skipped'])

        if self.stats['companies_no_matches']:
            report += "\n⚠️ EMPRESAS SIN MATCHES:\n"
            report += "\n".join(f"• {company}" for company in self.stats['companies_no_matches'])

        if self.stats['excluded_by_date_details']:
            report += f"\n🚫 XMLS EXCLUIDOS POR FECHA ({len(self.stats['excluded_by_date_details'])} total):\n"
            for excluded in self.stats['excluded_by_date_details'][:5]:  # Mostrar solo 5 ejemplos
                report += f"• {excluded['numero_consecutivo']} ({excluded['company']}) - {excluded['fecha_parsed']}\n"
            if len(self.stats['excluded_by_date_details']) > 5:
                remaining = len(self.stats['excluded_by_date_details']) - 5
                report += f"• ... y {remaining} más\n"

        if self.stats['excel_excluded_by_date_details']:
            report += f"\n🚫 FILAS EXCEL EXCLUIDAS POR FECHA ({len(self.stats['excel_excluded_by_date_details'])} total):\n"
            for excluded in self.stats['excel_excluded_by_date_details'][:5]:  # Mostrar solo 5 ejemplos
                report += f"• {excluded['numero_consecutivo']} ({excluded['filename']}) - {excluded['fecha_parsed']}\n"
            if len(self.stats['excel_excluded_by_date_details']) > 5:
                remaining = len(self.stats['excel_excluded_by_date_details']) - 5
                report += f"• ... y {remaining} más\n"

        if self.stats['correos_pdfs_processed'] > 0:
            report += f"\n📮 PROCESAMIENTO DE CORREOS:\n"
            report += f"• PDFs procesados exitosamente: {self.stats['correos_pdfs_processed']}\n"
            report += f"• PDFs con errores: {self.stats['correos_pdfs_failed']}\n"
            report += f"• Matches generados desde PDFs: {self.stats['correos_matches']}\n"

        if self.stats['otro_texto_processed'] > 0:
            report += f"\n🚗 EXTRACCIÓN DE PLACAS DESDE OTROTEXTO:\n"
            report += f"• Campos OtroTexto procesados: {self.stats['otro_texto_processed']}\n"
            report += f"• Placas extraídas exitosamente: {self.stats['placas_extracted']}\n"
            report += f"• Extracciones fallidas: {self.stats['placas_failed']}\n"
            report += f"• Fallback a Detalle utilizado: {self.stats['fallback_to_detalle']}\n"
            report += f"• Tasa de extracción: {self.stats['placa_extraction_rate']:.1f}%\n"

        return report